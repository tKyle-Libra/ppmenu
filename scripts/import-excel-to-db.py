#!/usr/bin/env python3

import zipfile
import xml.etree.ElementTree as ET
import openpyxl
import re
import os
import sys
import shutil
import sqlite3
import json
from io import BytesIO
from datetime import datetime
from PIL import Image

# ============================================================
# CONFIG
# ============================================================

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = os.path.join(BASE_DIR, '..', '价目表.xlsx')
DB_PATH = os.path.join(BASE_DIR, '..', 'ppmenu_assets_test', 'db', 'ppmenu.db')
OUTPUT_BASE = os.path.join(BASE_DIR, '..', 'ppmenu_assets_test')
CONFIG_PATH = os.path.join(BASE_DIR, 'src', 'utils', 'config.js')
PAGE_SIZE = 20

THUMBNAIL_SIZE = 300
JPEG_QUALITY = 90
THUMBNAIL_QUALITY = 80

SHEET_DIR_MAP = {
    '酱包': 'jb',
    '餐包': 'cb',
    '罐头': 'gt',
    '餐盒': 'ch',
    '奶制品': 'nzp',
    '汤': 'tang',
    '主食冻干': 'dg',
    '零食冻干': 'dg',
    '其他': 'other',
    '猫条': 'mt',
}

SHEET_TYPE_ID_MAP = {
    '酱包': 1,
    '餐包': 2,
    '罐头': 3,
    '餐盒': 4,
    '奶制品': 5,
    '汤': 6,
    '主食冻干': 7,
    '零食冻干': 7,
    '其他': 9,
    '猫条': 10,
}

BRAND_RENAME = {
    149: '阿飞与巴弟',
    11: 'Yee',
    10: 'URPet',
    148: '长胜饲养员',
    160: '鲜粮说',
}

NEW_BRANDS = ['GoMaoGo', '帕美德']

BRAND_ALIAS_MAP = {
    '阿飞与巴弟': 149,
    '阿飞和巴弟': 149,
    '阿飞巴弟': 149,
    'hachihachi': 3,
    'HACHIHACHI': 3,
    'Hachihachi': 3,
    'neku': 20,
    'Neku': 20,
    'Neku7': 20,
    'yee': 11,
    'YEE': 11,
    'Yee': 11,
    'urpet': 10,
    'URPET': 10,
    'URPet': 10,
    'likable': 17,
    'likable莱可宝': 17,
    'Likable莱可宝': 17,
    'likable莱可宝': 17,
    'GoMaoGo': 167,
    'GoMaoGao': 167,
    'GomaoGo': 167,
    'gomaaogo': 167,
    '幕野': 79,
    '慕野': 79,
    '我自有山海': 81,
    '我自由山海': 81,
    '森喜': 97,
    '森熹': 97,
    '维可派': 127,
    '维科派': 127,
    '柏丽高': 93,
    '佰丽高': 93,
    '它伢': 166,
    'tayaaa它伢': 166,
    'Tayaaa': 166,
    'tayaaa': 166,
    '喜崽': 59,
    '喜崽浓汤': 59,
    '季季予喵': 68,
    '季季予': 68,
    'ontutu': 22,
    'onTuTu': 22,
    'Ontutu': 22,
    'TheCat': 9,
    'TheCat': 9,
    'The Cat': 9,
    '巅峰': 31,
    'Ziwi巅峰': 31,
    '泡泡可儿': 102,
    'popocare': 102,
    'Popocare': 102,
    '大P便当': 63,
    '大P': 63,
    '帕美德': 168,
    '璞奇': 171,
    'wanli': 170,
}

EAT_TYPE_MAP = {
    '🐱': 1,
    '🐶': 2,
    '🐱🐶': 3,
    '🐶🐱': 3,
}

WEIGHT_RE = re.compile(r'(\d+(?:\.\d+)?\s*(?:g|ml|kg|L))')


def normalize_brand_key(name):
    return name.replace(' ', '').replace('\u3000', '').lower()


def build_brand_lookup(conn):
    cursor = conn.execute('SELECT brand_id, brand_name FROM brand')
    lookup = {}
    for bid, bname in cursor:
        lookup[normalize_brand_key(bname)] = bid
        lookup[bname] = bid
    for alias, bid in BRAND_ALIAS_MAP.items():
        if bid is not None:
            lookup[normalize_brand_key(alias)] = bid
            lookup[alias] = bid
    return lookup


def resolve_brand_id(brand_name, brand_lookup, conn):
    key = normalize_brand_key(brand_name)
    if key in brand_lookup:
        return brand_lookup[key]
    if brand_name in brand_lookup:
        return brand_lookup[brand_name]
    cursor = conn.execute(
        'SELECT brand_id FROM brand WHERE brand_name = ? COLLATE NOCASE',
        (brand_name,)
    )
    row = cursor.fetchone()
    if row:
        brand_lookup[key] = row[0]
        return row[0]
    cursor = conn.execute(
        'INSERT INTO brand (brand_name) VALUES (?)',
        (brand_name,)
    )
    new_id = cursor.lastrowid
    brand_lookup[key] = new_id
    brand_lookup[brand_name] = new_id
    print(f'  [新建品牌] {brand_name} (id={new_id})')
    return new_id


def parse_brand_col(value):
    if value is None:
        return None, None
    text = str(value)
    parts = text.split('\n')
    brand_name = parts[0].strip()
    series_parts = [p.strip() for p in parts[1:] if p.strip()]
    series_name = ' '.join(series_parts) if series_parts else ''
    return brand_name, series_name


def parse_eat_type(value):
    if value is None:
        return 1
    text = str(value).replace('\n', '').replace(' ', '')
    for key, etype in EAT_TYPE_MAP.items():
        if key in text:
            return etype
    if '猫' in text and '狗' in text:
        return 3
    if '狗' in text:
        return 2
    return 1


def parse_price(value):
    if value is None:
        return ''
    text = str(value).strip()
    text = text.rstrip('/')
    return text


def extract_weight(taste_text):
    if taste_text is None:
        return '', ''
    text = str(taste_text).strip()
    m = WEIGHT_RE.search(text)
    if m:
        weight = m.group(1).replace(' ', '')
        cleaned = text[:m.start()].strip().rstrip('，').rstrip(',').strip()
        cleaned = text[m.end():].strip() if m.end() < len(text) else ''
        remaining = text[:m.start()].strip()
        remaining = remaining.rstrip('，').rstrip(',').rstrip(' ').strip()
        return weight, remaining
    return '', text


PURE_WEIGHT_RE = re.compile(r'^[\d\s./～\-~gGmMlLkK]+$')


def extract_weight_v2(taste_text):
    if taste_text is None:
        return '', ''
    text = str(taste_text).strip()
    if PURE_WEIGHT_RE.match(text):
        return text.replace(' ', ''), ''
    matches = list(WEIGHT_RE.finditer(text))
    if matches:
        last = matches[-1]
        weight = last.group(1).replace(' ', '')
        remaining = text[:last.start()].strip()
        remaining = remaining.rstrip('，').rstrip(',').rstrip(' ').rstrip('\n').strip()
        after = text[last.end():].strip()
        if after and not PURE_WEIGHT_RE.match(after):
            if remaining:
                remaining = remaining + after
            else:
                remaining = after
        return weight, remaining
    return '', text


def clean_taste_name(name):
    name = name.strip().strip('\n').strip()
    m = WEIGHT_RE.match(name)
    if m and m.end() == len(name.replace(' ', '')):
        return ''
    m2 = re.match(r'^(.+?)\s*(\d+(?:\.\d+)?\s*[gGmMlL])$', name)
    if m2:
        return m2.group(1).strip()
    return name


def parse_tastes(taste_text):
    if taste_text is None or taste_text.strip() == '':
        return []
    text = taste_text.strip()
    if '、' in text:
        parts = text.split('、')
    elif '，' in text:
        parts = text.split('，')
    elif ',' in text:
        parts = text.split(',')
    else:
        cleaned = clean_taste_name(text)
        return [cleaned] if cleaned else []
    result = []
    for p in parts:
        cleaned = clean_taste_name(p)
        if cleaned:
            result.append(cleaned)
    return result


def check_is_young(product_name, taste_text, series_name):
    keywords = ['幼', '奶糕']
    combined = f"{product_name} {series_name} {taste_text or ''}"
    for kw in keywords:
        if kw in combined:
            return 1
    return 0


# ============================================================
# Step 1: 提取图片
# ============================================================

def extract_images(excel_path, output_base):
    print('\n=== Step 1: 提取图片 ===')

    if not os.path.exists(excel_path):
        print(f'[错误] Excel文件不存在: {excel_path}')
        return

    z = zipfile.ZipFile(excel_path)

    try:
        rels_content = z.read('xl/_rels/cellimages.xml.rels')
    except KeyError:
        print('[警告] 未找到 cellimages.xml.rels，跳过图片提取')
        z.close()
        return

    rels_root = ET.fromstring(rels_content)
    rid_map = {}
    for rel in rels_root:
        rid_map[rel.get('Id')] = 'xl/' + rel.get('Target').replace('../', '')

    ci_content = z.read('xl/cellimages.xml')
    ci_root = ET.fromstring(ci_content)

    ns1 = '{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}'
    ns2 = '{http://schemas.openxmlformats.org/drawingml/2006/main}'
    ns3 = '{http://schemas.openxmlformats.org/officeDocument/2006/relationships}'

    id_to_media = {}
    for cellImg in ci_root:
        pic = cellImg.find(f'{ns1}pic')
        if pic is None:
            continue
        nvPicPr = pic.find(f'{ns1}nvPicPr')
        cNvPr = nvPicPr.find(f'{ns1}cNvPr') if nvPicPr is not None else None
        if cNvPr is None:
            continue
        img_id = cNvPr.get('name', '')
        blipFill = pic.find(f'{ns1}blipFill')
        blip = blipFill.find(f'{ns2}blip') if blipFill is not None else None
        if blip is None:
            continue
        rid = blip.get(f'{ns3}embed', '')
        media_file = rid_map.get(rid, '')
        if img_id and media_file:
            id_to_media[img_id] = media_file

    print(f'[1/3] 解析到 {len(id_to_media)} 个图片映射')

    wb = openpyxl.load_workbook(excel_path, data_only=False)

    total_extracted = 0
    total_skipped = 0
    errors = []

    for sheet_name in wb.sheetnames:
        if sheet_name == 'WpsReserved_CellImgList':
            continue

        dir_name = SHEET_DIR_MAP.get(sheet_name, 'other')
        out_dir = os.path.join(output_base, dir_name)
        os.makedirs(out_dir, exist_ok=True)

        ws = wb[sheet_name]
        sheet_count = 0

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=6):
            cell_a = row[0]
            cell_f = row[5] if len(row) > 5 else None

            if not cell_a.value or not isinstance(cell_a.value, str) or 'DISPIMG' not in cell_a.value:
                continue

            match = re.search(r'DISPIMG\("([^"]+)"', cell_a.value)
            if not match:
                continue
            dispimg_id = match.group(1)

            if cell_f is None or cell_f.value is None or str(cell_f.value).strip() == '':
                continue
            filename = str(cell_f.value).strip()

            out_path = os.path.join(out_dir, filename + '.jpeg')

            if os.path.exists(out_path):
                total_skipped += 1
                continue

            media_path = id_to_media.get(dispimg_id)
            if media_path is None:
                errors.append(f'[{sheet_name}] 未找到图片: {dispimg_id}')
                continue

            try:
                img_data = z.read(media_path)
                img = Image.open(BytesIO(img_data))
                if img.mode in ('RGBA', 'P', 'LA'):
                    rgb = Image.new('RGB', img.size, (255, 255, 255))
                    if img.mode == 'P':
                        img = img.convert('RGBA')
                    rgb.paste(img, mask=img.split()[-1] if 'A' in img.mode else None)
                    img = rgb
                elif img.mode != 'RGB':
                    img = img.convert('RGB')

                img.save(out_path, 'JPEG', quality=THUMBNAIL_QUALITY)

                sheet_count += 1
            except Exception as e:
                errors.append(f'[{sheet_name}] {filename}: {e}')

        if sheet_count > 0:
            print(f'  [{sheet_name}] -> {dir_name}/ ({sheet_count} 张新提取)')
            total_extracted += sheet_count

    z.close()
    wb.close()

    print(f'[完成] 提取 {total_extracted} 张, 跳过已存在 {total_skipped} 张')
    if errors:
        print(f'[警告] {len(errors)} 个错误:')
        for e in errors[:10]:
            print(f'  {e}')


# ============================================================
# Step 2-5: 导入数据
# ============================================================

def import_data(excel_path, db_path):
    print('\n=== Step 2: 备份DB ===')
    if not os.path.exists(db_path):
        print(f'[错误] DB文件不存在: {db_path}')
        sys.exit(1)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_path = db_path + f'.bak.{timestamp}'
    shutil.copy2(db_path, backup_path)
    print(f'  备份到: {backup_path}')

    conn = sqlite3.connect(db_path)
    conn.execute('PRAGMA journal_mode=WAL')
    conn.row_factory = sqlite3.Row

    print('\n=== Step 3: 更新brand表 ===')
    for bid, new_name in BRAND_RENAME.items():
        conn.execute('UPDATE brand SET brand_name = ? WHERE brand_id = ?', (new_name, bid))
        print(f'  brand_id={bid} -> {new_name}')

    for bname in NEW_BRANDS:
        cursor = conn.execute('SELECT brand_id FROM brand WHERE brand_name = ?', (bname,))
        if cursor.fetchone() is None:
            conn.execute('INSERT INTO brand (brand_name) VALUES (?)', (bname,))
            print(f'  新建品牌: {bname}')

    conn.commit()

    print('\n=== Step 4: 构建brand映射表 ===')
    brand_lookup = build_brand_lookup(conn)
    print(f'  映射表条目: {len(brand_lookup)}')

    print('\n=== Step 5: 遍历Sheet导入数据 ===')

    wb = openpyxl.load_workbook(excel_path, data_only=True)

    stats = {
        'products_inserted': 0,
        'products_updated': 0,
        'tastes_inserted': 0,
        'errors': [],
    }

    for sheet_name in wb.sheetnames:
        if sheet_name == 'WpsReserved_CellImgList':
            continue

        type_id = SHEET_TYPE_ID_MAP.get(sheet_name)
        if type_id is None:
            print(f'  [{sheet_name}] 未知Sheet，跳过')
            continue

        dir_prefix = SHEET_DIR_MAP.get(sheet_name, 'other')
        is_snacks = 1 if sheet_name == '零食冻干' else 0

        ws = wb[sheet_name]
        sheet_inserted = 0
        sheet_updated = 0

        for row_idx, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=6), start=3):
            col_brand = row[1].value if len(row) > 1 else None
            col_taste = row[2].value if len(row) > 2 else None
            col_type = row[3].value if len(row) > 3 else None
            col_price = row[4].value if len(row) > 4 else None
            col_img = row[5].value if len(row) > 5 else None

            if col_brand is None and col_taste is None:
                continue

            brand_name, series_name = parse_brand_col(col_brand)
            if brand_name is None or brand_name.strip() == '':
                continue

            brand_id = resolve_brand_id(brand_name, brand_lookup, conn)
            conn.commit()

            product_name = series_name if series_name else brand_name

            weight, taste_text = extract_weight_v2(col_taste)

            price = parse_price(col_price)

            eat_type = parse_eat_type(col_type)

            if col_img is not None and str(col_img).strip():
                img_name = str(col_img).strip()
                if not img_name.lower().endswith(('.jpeg', '.jpg', '.png')):
                    img_name += '.jpeg'
                product_img = f'{dir_prefix}/{img_name}'
            else:
                product_img = 'default.png'

            is_young_global = check_is_young(product_name, col_taste, series_name)

            product_id = None
            operation = None

            cursor = conn.execute(
                'SELECT product_id FROM product WHERE product_name = ? AND brand_id = ? AND product_type_id = ? AND product_weight = ?',
                (product_name, brand_id, type_id, weight)
            )
            existing = cursor.fetchone()

            if existing:
                product_id = existing[0]
                conn.execute('''
                    UPDATE product SET
                        product_price = ?,
                        product_img = ?,
                        product_eat_type = ?,
                        product_is_new = 0
                    WHERE product_id = ?
                ''', (price, product_img, eat_type, product_id))
                operation = 'updated'
                sheet_updated += 1
                stats['products_updated'] += 1
            else:
                if weight:
                    cursor2 = conn.execute(
                        'SELECT product_id FROM product WHERE product_name = ? AND brand_id = ? AND product_type_id = ? AND product_weight = ?',
                        (product_name, brand_id, type_id, '')
                    )
                    existing2 = cursor2.fetchone()
                    if existing2:
                        product_id = existing2[0]
                        conn.execute('''
                            UPDATE product SET
                                product_price = ?,
                                product_img = ?,
                                product_weight = ?,
                                product_eat_type = ?,
                                product_is_new = 0
                            WHERE product_id = ?
                        ''', (price, product_img, weight, eat_type, product_id))
                        operation = 'updated'
                        sheet_updated += 1
                        stats['products_updated'] += 1

                if product_id is None:
                    conn.execute('''
                        INSERT INTO product (product_name, product_price, brand_id, product_img, product_weight, product_eat_type, product_is_new, product_type_id)
                        VALUES (?, ?, ?, ?, ?, ?, 0, ?)
                    ''', (product_name, price, brand_id, product_img, weight, eat_type, type_id))
                    product_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
                    operation = 'inserted'
                    sheet_inserted += 1
                    stats['products_inserted'] += 1

            conn.execute('DELETE FROM product_taste WHERE product_id = ?', (product_id,))

            tastes = parse_tastes(taste_text)
            for taste_name in tastes:
                is_young = is_young_global
                if is_young == 0:
                    is_young = check_is_young('', taste_name, '')
                try:
                    conn.execute('''
                        INSERT INTO product_taste (product_taste_name, product_id, is_young, is_snacks)
                        VALUES (?, ?, ?, ?)
                    ''', (taste_name, product_id, is_young, is_snacks))
                    stats['tastes_inserted'] += 1
                except Exception as e:
                    stats['errors'].append(f'[{sheet_name}:row{row_idx}] taste "{taste_name}" for product_id={product_id}: {e}')

            conn.commit()

        action_parts = []
        if sheet_inserted:
            action_parts.append(f'新增{sheet_inserted}')
        if sheet_updated:
            action_parts.append(f'更新{sheet_updated}')
        action_str = ', '.join(action_parts) if action_parts else '无变化'
        print(f'  [{sheet_name}] {action_str}')

    wb.close()

    print('\n=== Step 6: 清理重复产品 ===')
    cleanup_duplicates(conn)

    conn.close()

    print(f'\n=== Step 7: 统计 ===')
    print(f'  新增产品: {stats["products_inserted"]}')
    print(f'  更新产品: {stats["products_updated"]}')
    print(f'  新增口味: {stats["tastes_inserted"]}')
    if stats['errors']:
        print(f'  错误: {len(stats["errors"])}')
        for e in stats['errors'][:20]:
            print(f'    {e}')


def cleanup_duplicates(conn):
    cur = conn.execute('''
        SELECT product_name, brand_id, product_type_id, COUNT(*) as cnt,
               GROUP_CONCAT(product_id) as ids,
               GROUP_CONCAT(product_weight) as weights
        FROM product
        GROUP BY product_name, brand_id, product_type_id
        HAVING cnt > 1
    ''')
    groups = cur.fetchall()
    if not groups:
        print('  无重复产品')
        return

    deleted_count = 0
    for group in groups:
        name, bid, tid, cnt, ids_str, weights_str = group
        ids = ids_str.split(',')
        weights = weights_str.split(',')

        keep_id = None
        remove_ids = []

        empty_weight_ids = []
        has_weight_ids = []
        for pid, w in zip(ids, weights):
            if w:
                has_weight_ids.append((pid, w))
            else:
                empty_weight_ids.append(pid)

        if has_weight_ids:
            keep_id = has_weight_ids[0][0]
            remove_ids = [pid for pid, _ in has_weight_ids[1:]]
            remove_ids.extend(empty_weight_ids)
        else:
            keep_id = ids[0]
            remove_ids = ids[1:]

        if remove_ids:
            for pid in remove_ids:
                conn.execute('DELETE FROM product_taste WHERE product_id = ?', (int(pid),))
                conn.execute('DELETE FROM product WHERE product_id = ?', (int(pid),))
                deleted_count += 1

    conn.commit()
    print(f'  清理 {deleted_count} 条重复产品（合并了 {len(groups)} 组）')


def regenerate_json(db_path):
    db_dir = os.path.dirname(db_path)
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row

    cur = conn.execute('''
        SELECT 
            p.product_id as id,
            p.product_name as name,
            p.product_price as price,
            p.product_img as image,
            p.product_weight as weight,
            p.product_eat_type as eatType,
            p.product_is_new as isNew,
            p.product_type_id as typeId,
            b.brand_name as brand,
            GROUP_CONCAT(pt.product_taste_name, '、') as tastes,
            MAX(pt.is_young) as is_young
        FROM product p
        JOIN brand b ON p.brand_id = b.brand_id
        LEFT JOIN product_taste pt ON p.product_id = pt.product_id
        GROUP BY p.product_id
        ORDER BY p.product_id
    ''')

    products = []
    for row in cur:
        product = {
            'id': row['id'],
            'name': row['name'],
            'price': row['price'],
            'image': row['image'],
            'weight': row['weight'],
            'eatType': row['eatType'],
            'isNew': row['isNew'] == 1,
            'is_young': row['is_young'] == 1,
            'typeId': row['typeId'],
            'brand': row['brand'],
            'tastes': row['tastes'] or '',
        }
        products.append(product)

    total = len(products)
    total_pages = (total + PAGE_SIZE - 1) // PAGE_SIZE
    conn.close()

    for f in os.listdir(db_dir):
        if f.startswith('data-page') and f.endswith('.json'):
            os.remove(os.path.join(db_dir, f))

    for page in range(1, total_pages + 1):
        start = (page - 1) * PAGE_SIZE
        end = start + PAGE_SIZE
        page_data = products[start:end]
        filepath = os.path.join(db_dir, f'data-page{page}.json')
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(page_data, f, ensure_ascii=False, indent=2)

    print(f'  生成 {total_pages} 页 JSON, 共 {total} 条产品')
    return total_pages


def update_config(db_path):
    print('\n=== 更新 config.js ===')

    total_pages = regenerate_json(db_path)

    config_path = os.path.abspath(CONFIG_PATH)
    if not os.path.exists(config_path):
        print(f'  [跳过] config.js 不存在: {config_path}')
        return

    with open(config_path, 'r', encoding='utf-8') as f:
        content = f.read()

    new_pages_str = str(total_pages)
    content = re.sub(
        r"(TOTAL_PAGES:\s*)\d+",
        rf"\g<1>{new_pages_str}",
        content,
    )

    with open(config_path, 'w', encoding='utf-8') as f:
        f.write(content)

    print(f'  TOTAL_PAGES -> {new_pages_str}')


def main():
    excel_path = os.path.abspath(EXCEL_PATH)
    db_path = os.path.abspath(DB_PATH)
    output_base = os.path.abspath(OUTPUT_BASE)

    print('=== Excel价目表 → DB导入工具 ===')
    print(f'Excel: {excel_path}')
    print(f'DB:    {db_path}')
    print(f'图片:  {output_base}')

    if not os.path.exists(excel_path):
        print(f'[错误] Excel文件不存在: {excel_path}')
        sys.exit(1)

    extract_images(excel_path, output_base)
    import_data(excel_path, db_path)
    update_config(db_path)

    print('\n✅ 全部完成')


if __name__ == '__main__':
    main()
